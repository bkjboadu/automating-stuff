import openpyxl,sys
from openpyxl.utils import get_column_letter

if len(sys.argv) > 1:
    a = int(sys.argv[1])
else:
    a = 6

new = openpyxl.Workbook()
sheet = new.active
while a >= 1:
    sheet.cell(row=1,column=a + 1).value = a
    sheet.cell(row=a + 1,column = 1).value = a
    a -= 1

for row in range(2,sheet.max_row + 1):
    for column in range(2,sheet.max_column + 1):
        sheet.cell(row=row,column=column).value = f'=A{row}*{get_column_letter(column)}1'
new.save('multi.xlsx')


