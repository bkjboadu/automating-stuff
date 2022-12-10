import sys
import openpyxl

if len(sys.argv) > 1:
    filename = sys.argv[1]
else:
    filename = 'work.xlsx'

wb = openpyxl.load_workbook(filename)
wba = wb.active

list = []
for i in range(wba.max_column):
    list.append([])

for x in range(len(list)):
    for i in range(wba.max_row):
        list[x].append(wba.cell(row=i+1,column=x+1).value)


new = openpyxl.Workbook()
new_a = new.active
for x in range(len(list)):
    for i in range(wba.max_row):
        new_a.cell(row=x+1,column=i+1).value = list[x][i]


new.save('inverted.xlsx')
