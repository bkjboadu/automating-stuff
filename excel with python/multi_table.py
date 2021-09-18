import sys,openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

number = int(sys.argv[1])
cells = number + 1

wb = openpyxl.Workbook()
sheet = wb.active

make_bold = Font(bold=True)

while cells > 1:
    sheet.cell(row=cells,column=1).value = cells - 1
    sheet.cell(row=1,column=cells).value = cells - 1

    sheet.cell(row=cells,column=1).font = make_bold
    sheet.cell(row=1,column=cells).font = make_bold

    cells -= 1

in_column = sheet.max_column

while in_column > 1:
    row_numbers = sheet.max_row
    column_header = get_column_letter(in_column)
    while row_numbers > 1:
        sheet[column_header + str(row_numbers)] = '=(' + column_header + '1*A' + str(row_numbers) +')'
        row_numbers -= 1
    in_column -= 1



wb.save('multi_table.xlsx')
