import openpyxl,sys

if len(sys.argv) > 2:
    filename = sys.argv[1]
else:
    filename = 'done.xlsx'

wb = openpyxl.load_workbook(filename)
wba = wb.active


for i in range(wba.max_column):
    text_name = f"text{i+1}.txt"
    with open(text_name,'w') as f:
        for x in range(wba.max_row):
            if wba.cell(row=x+1,column=i+1).value == None:
                continue
            else:
                f.write(wba.cell(row=x+1,column=i+1).value)

