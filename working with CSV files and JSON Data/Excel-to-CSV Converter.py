import openpyxl,csv,os
from pathlib import Path

for file in Path.cwd().glob('*.xlsx'):
    wb = openpyxl.load_workbook(file)
    for sheet in wb.sheetnames:
        excel_data =  []
        active_wb = wb[sheet]
        for rows in range(active_wb.max_row):
            excel_data.append([])
            for columns in range(active_wb.max_column):
                excel_data[rows].append(active_wb.cell(row=rows+1,column=columns+1).value)

        filename = str(Path(file).stem) + '_' + active_wb.title + '.csv'
        new_csv = open(filename,'w',newline='')
        csvwriter = csv.writer(new_csv)
        for row in excel_data:
            csvwriter.writerow(row)
        new_csv.close()



